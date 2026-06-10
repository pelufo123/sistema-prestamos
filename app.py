import os
import io
import psycopg2

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    jsonify,
    send_file
)

from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

app = Flask(__name__)

app.secret_key = "clave_super_segura"


# ------------------------------
# 🔐 PROTEGER RUTAS
# ------------------------------
@app.before_request
def proteger_rutas():

    rutas_libres = [
        "login",
        "static"
    ]

    # 🔥 evitar errores
    if request.endpoint is None:
        return

    # 🔥 permitir login y archivos static
    if request.endpoint in rutas_libres:
        return

    # 🔥 si NO hay sesión → login
    if "usuario" not in session:
        return redirect(url_for("login"))
# ------------------------------
# 🔌 CONEXIÓN A BASE DE DATOS
# ------------------------------
def conectar():
    db_url = os.getenv("DATABASE_URL")

    if not db_url:
        print("❌ No hay DATABASE_URL")
        return None

    db_url = db_url.replace("postgres://", "postgresql://", 1)

    try:
        conn = psycopg2.connect(
            db_url,
            sslmode="require"
        )
        return conn

    except Exception as e:
        print("❌ Error conexión DB:", e)
        return None
#323 2879658
# ------------------------------
def init_db():

    conn = conectar()

    if not conn:
        print("❌ No hay conexión a la base de datos")
        return

    cur = conn.cursor()

    try:

        # =====================================
        # 👤 USUARIOS
        # =====================================
        cur.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (

            id SERIAL PRIMARY KEY,

            username TEXT UNIQUE,

            password TEXT

        );
        """)

        # =====================================
        # 👥 CLIENTES
        # =====================================
        cur.execute("""
        CREATE TABLE IF NOT EXISTS clientes (

            id SERIAL PRIMARY KEY,

            nombre TEXT,

            telefono TEXT,

            direccion TEXT,

            usuario TEXT

        );
        """)

        # =====================================
        # 💰 PRESTAMOS
        # =====================================
        cur.execute("""
        CREATE TABLE IF NOT EXISTS prestamos(

            id SERIAL PRIMARY KEY,

            cliente_id INTEGER REFERENCES clientes(id),

            capital REAL,

            interes REAL,

            dias INTEGER,

            fecha DATE,

            vencimiento DATE,

            total REAL,

            estado TEXT DEFAULT 'Activo',

            usuario TEXT

        );
        """)

        # =====================================
        # 🔥 AGREGAR COLUMNA ESTADO
        # SI NO EXISTE
        # =====================================
        try:

            cur.execute("""
                ALTER TABLE prestamos
                ADD COLUMN estado TEXT DEFAULT 'Activo'
            """)

            conn.commit()

            print("✅ Columna estado agregada")

        except Exception as e:

            print("ℹ️ La columna estado ya existe")

        # =====================================
        # 💸 ABONOS
        # =====================================
        cur.execute("""
        CREATE TABLE IF NOT EXISTS abonos(

            id SERIAL PRIMARY KEY,

            prestamo_id INTEGER REFERENCES prestamos(id),

            monto REAL,

            fecha TIMESTAMP,

            tipo TEXT,

            mes INTEGER,

            usuario TEXT

        );
        """)

        # =====================================
        # 🏦 CAJA
        # =====================================
        cur.execute("""
        CREATE TABLE IF NOT EXISTS caja(

            id SERIAL PRIMARY KEY,

            tipo TEXT,

            monto REAL,

            descripcion TEXT,

            fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP

        );
        """)

        conn.commit()

        print("✅ Base de datos lista")

    except Exception as e:

        print("❌ Error en DB:", e)

    finally:

        conn.close()
def crear_admin():
    conn = conectar()

    if not conn:
        return

    cur = conn.cursor()

    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id SERIAL PRIMARY KEY,
                username VARCHAR(100) UNIQUE NOT NULL,
                password VARCHAR(100) NOT NULL
            )
        """)

        conn.commit()

        cur.execute(
            "SELECT * FROM usuarios WHERE username=%s",
            ("admin",)
        )

        if not cur.fetchone():
            cur.execute(
                "INSERT INTO usuarios (username, password) VALUES (%s, %s)",
                ("admin", "1234")
            )
            conn.commit()

    except Exception as e:
        print("Error:", e)

    finally:
        conn.close()
# ------------------------------
# 💲 FORMATO
# ------------------------------
def formato(x):
    try:
        return "{:,.0f}".format(x).replace(",", ".")
    except:
        return "0"
# ------------------------------
# 📆 MESES DE ATRASO
# ------------------------------
def meses_atraso(fecha_inicio):
    hoy = datetime.now().date()

    if isinstance(fecha_inicio, str):
        fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()

    return (hoy.year - fecha_inicio.year) * 12 + (hoy.month - fecha_inicio.month)
# ------------------------------
# 🔹 VALIDACIÓN
# ------------------------------
def cliente_valido(cliente_id):
    return cliente_id and str(cliente_id).isdigit()

# ------------------------------
def interes_acumulado(pid, conn):
    cur = conn.cursor()

    cur.execute("SELECT capital, interes, fecha FROM prestamos WHERE id=%s", (pid,))
    data = cur.fetchone()

    if not data:
        return 0, 0

    capital, interes, fecha = data

    # 🔥 meses reales transcurridos
    meses = meses_atraso(fecha)

    if meses < 1:
        return 0, 0

    # 🔥 interés total generado
    interes_total = capital * (interes / 100) * meses

    # 🔥 interés ya pagado
    cur.execute("""
        SELECT SUM(monto)
        FROM abonos
        WHERE prestamo_id=%s AND tipo='interes'
    """, (pid,))
    pagado = cur.fetchone()[0] or 0

    deuda = interes_total - pagado

    if deuda < 0:
        deuda = 0

    return deuda, meses
# ------------------------------
# 🔥 INTERÉS DIARIO
# ------------------------------
def interes_hoy(pid, conn):
    cur = conn.cursor()
    hoy = datetime.now().date()

    cur.execute("""
        SELECT COUNT(*) FROM abonos
        WHERE prestamo_id=%s AND tipo='interes' AND DATE(fecha)=%s
    """, (pid, hoy))

    pago_hoy = cur.fetchone()[0]

    cur.execute("SELECT capital, interes FROM prestamos WHERE id=%s", (pid,))
    data = cur.fetchone()

    if not data:
        return 0

    capital, interes = data
    interes_total = capital * interes / 100

    return 0 if pago_hoy > 0 else interes_total

# ------------------------------
# ------------------------------
# ------------------------------
# 🧠 CÁLCULO
def calcular(pid, conn):

    cur = conn.cursor()

    # =====================================
    # DATOS PRÉSTAMO
    # =====================================
    cur.execute("""
        SELECT capital, interes, fecha
        FROM prestamos
        WHERE id=%s
    """, (pid,))

    data = cur.fetchone()

    if not data:
        return 0, 0, 0, 0, 0

    capital_original = float(data[0])
    interes = float(data[1])
    fecha = data[2]

    # =====================================
    # CAPITAL ABONADO
    # =====================================
    cur.execute("""
        SELECT COALESCE(SUM(monto),0)
        FROM abonos
        WHERE prestamo_id=%s
        AND tipo='capital'
    """, (pid,))

    abonado_capital = float(
        cur.fetchone()[0] or 0
    )

    capital_restante = (
        capital_original - abonado_capital
    )

    if capital_restante < 0:
        capital_restante = 0

    # =====================================
    # MESES TRANSCURRIDOS
    # =====================================
    meses = meses_atraso(fecha)

    if meses < 1:
        meses = 1

    # =====================================
    # INTERÉS MENSUAL
    # =====================================
    interes_mensual = (
        capital_restante *
        (interes / 100)
    )

    # =====================================
    # INTERÉS GENERADO
    # =====================================
    interes_generado = (
        interes_mensual * meses
    )

    # =====================================
    # INTERÉS PAGADO
    # =====================================
    cur.execute("""
        SELECT COALESCE(SUM(monto),0)
        FROM abonos
        WHERE prestamo_id=%s
        AND tipo='interes'
    """, (pid,))

    abonado_interes = float(
        cur.fetchone()[0] or 0
    )

    interes_restante = (
        interes_generado -
        abonado_interes
    )

    if interes_restante < 0:
        interes_restante = 0

    # =====================================
    # TOTAL
    # =====================================
    saldo_total = (
        capital_restante +
        interes_restante
    )

    return (

        round(capital_restante),
        round(interes_restante),
        round(saldo_total),
        round(abonado_capital),
        round(abonado_interes)

    )
# ------------------------------
# 🏠 INICIO
# ------------------------------
@app.route("/")
def inicio():

    # 🔥 si NO ha iniciado sesión
    if "usuario" not in session:
        return redirect(url_for("login"))

    # 🔥 si ya inició sesión
    return redirect(url_for("panel"))
# =============================
# 👥 GANANCIA POR CLIENTE
# =============================
def ganancia_por_cliente(conn):

    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre
        FROM clientes
    """)

    resultado = []

    clientes = cur.fetchall()

    for cliente in clientes:

        cid = cliente[0]
        nombre = cliente[1]

        # 🔥 TOTAL PRESTADO
        cur.execute("""
            SELECT SUM(capital)
            FROM prestamos
            WHERE cliente_id=%s
        """, (cid,))

        capital = cur.fetchone()[0] or 0

        # 🔥 TOTAL CAPITAL RECUPERADO
        cur.execute("""
            SELECT SUM(a.monto)
            FROM abonos a
            JOIN prestamos p
            ON a.prestamo_id = p.id
            WHERE p.cliente_id=%s
            AND a.tipo='capital'
        """, (cid,))

        recuperado = cur.fetchone()[0] or 0

        # 🔥 TOTAL INTERÉS GANADO
        cur.execute("""
            SELECT SUM(a.monto)
            FROM abonos a
            JOIN prestamos p
            ON a.prestamo_id = p.id
            WHERE p.cliente_id=%s
            AND a.tipo='interes'
        """, (cid,))

        interes = cur.fetchone()[0] or 0

        deuda = capital - recuperado

        resultado.append({
            "nombre": nombre,
            "capital": capital,
            "recuperado": recuperado,
            "interes": interes,
            "deuda": deuda
        })

    return resultado
# ------------------------------
# 📊 PANEL PRINCIPAL
# ------------------------------
@app.route("/panel", methods=["GET", "POST"])
def panel():

    conn = conectar()

    if conn is None:
        return "Error de conexión", 500

    cur = conn.cursor()

    # =============================
    # 🔥 FILTRO FECHA
    # =============================
    tipo = request.form.get("tipo") or "dia"

    fecha_str = request.form.get("fecha")

    if fecha_str:

        fecha = datetime.strptime(
            fecha_str,
            "%Y-%m-%d"
        ).date()

    else:

        fecha = datetime.now().date()

    # =============================
    # 💰 CAPITAL EN CALLE
    # =============================
    capital_total = 0

    cur.execute("SELECT id FROM prestamos")

    for (pid,) in cur.fetchall():

        cap_rest, _, _, _, _ = calcular(pid, conn)

        if cap_rest > 0:

            capital_total += cap_rest

    # =============================
    # 📅 CAPITAL E INTERÉS DEL DÍA
    # =============================
    cur.execute("""
        SELECT monto, tipo
        FROM abonos
        WHERE DATE(fecha) = %s
    """, (fecha,))

    capital_dia = 0
    interes_dia = 0

    for monto, tipo_pago in cur.fetchall():

        if tipo_pago == "capital":

            capital_dia += monto

        else:

            interes_dia += monto

    # =============================
    # 📆 CAPITAL E INTERÉS MES
    # =============================
    inicio_mes = fecha.replace(day=1)

    cur.execute("""
        SELECT monto, tipo
        FROM abonos
        WHERE DATE(fecha) >= %s
        AND DATE(fecha) <= %s
    """, (inicio_mes, fecha))

    capital_mes = 0
    interes_mes = 0

    for monto, tipo_pago in cur.fetchall():

        if tipo_pago == "capital":

            capital_mes += monto

        else:

            interes_mes += monto


    # =============================
    # ⚠️ PRÓXIMOS Y VENCIDOS
    # =============================
    cur.execute("""
        SELECT
            p.id,
            p.fecha,
            c.nombre

        FROM prestamos p

        JOIN clientes c
        ON p.cliente_id = c.id
    """)

    por_vencer = []
    vencidos = []

    hoy = datetime.now().date()

    for pid, fecha_inicio, nombre in cur.fetchall():

        # 🔥 saldo real actualizado
        cap_rest, int_rest, _, _, _ = calcular(pid, conn)

        saldo = cap_rest + int_rest

        # 🔥 si ya terminó
        if saldo <= 0:
            continue

        # 🔥 vencimiento automático 30 días
        fecha_vencimiento = fecha_inicio + timedelta(days=30)

        # 🔥 días restantes
        dias_restantes = (
            fecha_vencimiento - hoy
        ).days

        # =============================
        # 🔴 VENCIDOS
        # =============================
        if dias_restantes < 0:

            dias_atraso = abs(dias_restantes)

            vencidos.append((
                pid,
                nombre,
                dias_atraso,
                formato(saldo)
            ))

        # =============================
        # 🟠 PRÓXIMOS A VENCER
        # =============================
        elif dias_restantes <= 5:

            por_vencer.append((
                pid,
                nombre,
                dias_restantes,
                formato(saldo)
            ))

    # =============================
    # 👥 GANANCIA CLIENTES
    # =============================
    clientes_ganancia = ganancia_por_cliente(conn)

    # =============================
    # 📋 PRÉSTAMOS ACTIVOS
    # =============================
    cur.execute("""
        SELECT
            p.id,
            c.nombre,
            p.capital,
            p.interes

        FROM prestamos p

        JOIN clientes c
        ON p.cliente_id = c.id
    """)

    prestamos = []

    rows = cur.fetchall()

    for pid, nombre, capital, interes in rows:

        # =========================
        # CÁLCULO
        # =========================
        capital_restante, interes_restante, saldo_total, _, _ = calcular(pid, conn)

        # =========================
        # FECHA DEL PRÉSTAMOS
        # =========================
        cur.execute("""
            SELECT fecha
            FROM prestamos
            WHERE id=%s
        """, (pid,))

        fecha_prestamo = cur.fetchone()[0]

        # =========================
        # MESES PAGADOS
        # =========================
        cur.execute("""
            SELECT mes
            FROM abonos
            WHERE prestamo_id=%s
            AND tipo='interes'
            ORDER BY mes ASC
        """, (pid,))

        meses_pagados = [x[0] for x in cur.fetchall()]

        # =========================
        # GENERAR 24 MESES
        # =========================
        meses = []

        fecha_base = fecha_prestamo

        for i in range(24):

            mes_num = i + 1

            nueva_fecha = fecha_base + relativedelta(months=i)

            texto = nueva_fecha.strftime("%B %Y").capitalize()

            meses.append({
                "numero": mes_num,
                "texto": texto
            })

        # =========================
        # SIGUIENTE MES PENDIENTE
        # =========================
        siguiente_mes = 1

        while siguiente_mes in meses_pagados:
            siguiente_mes += 1

        # =========================
        # TEXTO DEL SIGUIENTE MES
        # =========================
        siguiente_mes_texto = ""

        for m in meses:
            if m["numero"] == siguiente_mes:
                siguiente_mes_texto = m["texto"]

        # =========================
        # ÚLTIMO MES PAGADO
        # =========================
        ultimo_pagado = ""

        if meses_pagados:

            ultimo_num = max(meses_pagados)

            for m in meses:
                if m["numero"] == ultimo_num:
                    ultimo_pagado = m["texto"]

        # =========================
        # GUARDAR
        # =========================
        prestamos.append({

            "id": pid,
            "nombre": nombre,

            "capital": formato(capital_restante),
            "interes": formato(interes_restante),
            "total": formato(saldo_total),

            "interes_num": interes_restante,

            "meses": meses,

            "siguiente_mes": siguiente_mes,
            "siguiente_mes_texto": siguiente_mes_texto,

            "ultimo_pagado": ultimo_pagado
        })

    # =============================
    # 🔧 VARIABLES
    # =============================
    capital = capital_dia
    interes = interes_dia

    cur.close()
    conn.close()

    return render_template(

        "panel.html",

        fecha=fecha,

        tipo=tipo,

        clientes_ganancia=clientes_ganancia,

        capital=formato(capital),

        interes=formato(interes),

        capital_total=formato(capital_total),

        capital_dia=formato(capital_dia),

        interes_dia=formato(interes_dia),

        capital_mes=formato(capital_mes),

        interes_mes=formato(interes_mes),

        por_vencer=por_vencer,

        vencidos=vencidos,

        prestamos=prestamos
)

def crear_tabla_caja():
    conn = conectar()

    if not conn:
        return

    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS caja (
            id SERIAL PRIMARY KEY,
            tipo VARCHAR(20) NOT NULL,
            monto NUMERIC(15,2) NOT NULL,
            descripcion VARCHAR(255),
            fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.commit()
    conn.close()

#  CAJA
@app.route("/caja", methods=["GET", "POST"])
def caja():

    conn = conectar()

    if not conn:
        return "Error conexión DB"

    cur = conn.cursor()

    error = ""
    mensaje = ""

    # ====================================
    # 💾 GUARDAR NUEVO MOVIMIENTO
    # ====================================
    if request.method == "POST":

        try:

            tipo = request.form["tipo"]

            monto = float(
                request.form["monto"]
                .replace(".", "")
                .replace(",", "")
            )

            descripcion = request.form["descripcion"]

            # ====================================
            # 💰 CALCULAR DISPONIBLE ACTUAL
            # ====================================
            cur.execute("""
                SELECT

                    COALESCE(SUM(
                        CASE
                            WHEN tipo='ingreso'
                            THEN monto
                            ELSE 0
                        END
                    ),0),

                    COALESCE(SUM(
                        CASE
                            WHEN tipo='egreso'
                            THEN monto
                            ELSE 0
                        END
                    ),0)

                FROM caja
            """)

            datos = cur.fetchone()

            ingresos_actuales = float(
                datos[0] or 0
            )

            egresos_actuales = float(
                datos[1] or 0
            )

            disponible_actual = (
                ingresos_actuales -
                egresos_actuales
            )

            # ====================================
            # ❌ VALIDAR FONDOS
            # ====================================
            if (
                tipo == "egreso"
                and monto > disponible_actual
            ):

                error = (
                    f"❌ No hay suficiente dinero en caja. "
                    f"Disponible: ${formato(disponible_actual)}"
                )

            else:

                # ====================================
                # 💾 GUARDAR
                # ====================================
                cur.execute("""
                    INSERT INTO caja (

                        tipo,
                        monto,
                        descripcion

                    )
                    VALUES (%s,%s,%s)
                """, (

                    tipo,
                    monto,
                    descripcion

                ))

                conn.commit()

                mensaje = "✅ Movimiento guardado"

        except Exception as e:

            print("ERROR CAJA:", e)

            error = "❌ Error guardando movimiento"

    # ====================================
    # 📅 FILTRO POR MES
    # ====================================
    mes = request.args.get("mes")

    if mes:

        fecha = datetime.strptime(
            mes,
            "%Y-%m"
        )

    else:

        fecha = datetime.now()

    # ====================================
    # 🔥 INICIO MES
    # ====================================
    inicio_mes = fecha.replace(

        day=1,
        hour=0,
        minute=0,
        second=0,
        microsecond=0

    )

    # ====================================
    # 🔥 SIGUIENTE MES
    # ====================================
    if fecha.month == 12:

        siguiente = fecha.replace(

            year=fecha.year + 1,
            month=1,
            day=1,
            hour=0,
            minute=0,
            second=0,
            microsecond=0

        )

    else:

        siguiente = fecha.replace(

            month=fecha.month + 1,
            day=1,
            hour=0,
            minute=0,
            second=0,
            microsecond=0

        )

    # ====================================
    # 📋 MOVIMIENTOS DEL MES
    # ====================================
    cur.execute("""
        SELECT

            id,
            tipo,
            monto,
            descripcion,
            fecha

        FROM caja

        WHERE fecha >= %s
        AND fecha < %s

        ORDER BY id DESC

    """, (

        inicio_mes,
        siguiente

    ))

    movimientos = cur.fetchall()

    # ====================================
    # 🔥 INICIAR TOTALES EN 0
    # ====================================
    ingresos = 0
    egresos = 0
    disponible = 0

    # ====================================
    # 🔥 SUMAR MOVIMIENTOS DEL MES
    # ====================================
    for mov in movimientos:

        monto_mov = float(mov[2])

        if mov[1] == "ingreso":

            ingresos += monto_mov

        else:

            egresos += monto_mov

    # ====================================
    # 💰 DISPONIBLE
    # ====================================
    disponible = ingresos - egresos

    conn.close()

    return render_template(

        "caja.html",

        movimientos=movimientos,

        ingresos=ingresos,

        egresos=egresos,

        disponible=disponible,

        formato=formato,

        mes=fecha.strftime("%Y-%m"),

        error=error,

        mensaje=mensaje
    )

# ====================================
# 📊 REPORTE GENERAL
# ====================================
@app.route("/reporte_general")
def reporte_general():

    conn = conectar()

    if not conn:
        return "Error conexión DB"

    cur = conn.cursor()

    # ====================================
    # 📅 FILTRO MES
    # ====================================
    mes = request.args.get("mes")

    if mes:

        fecha = datetime.strptime(
            mes,
            "%Y-%m"
        )

    else:

        fecha = datetime.now()

    inicio_mes = fecha.replace(
        day=1
    ).date()

    # 🔥 siguiente mes
    if fecha.month == 12:

        siguiente = fecha.replace(
            year=fecha.year + 1,
            month=1,
            day=1
        ).date()

    else:

        siguiente = fecha.replace(
            month=fecha.month + 1,
            day=1
        ).date()

    # ====================================
    # 👥 CLIENTES
    # ====================================
    cur.execute("""
        SELECT
            id,
            nombre,
            telefono
        FROM clientes
        ORDER BY nombre ASC
    """)

    clientes = cur.fetchall()

    reporte = []

    # ====================================
    # 🔥 RECORRER CLIENTES
    # ====================================
    for cliente in clientes:

        cliente_id = cliente[0]

        # ====================================
        # 📋 PRÉSTAMOS
        # ====================================
        cur.execute("""
                SELECT
                    id,
                    capital,
                    interes,
                    fecha,
                    total,
                    estado
                FROM prestamos
                WHERE cliente_id=%s
        """, (cliente_id,))

        prestamos = cur.fetchall()

        total_prestado = 0
        saldo_total = 0
        intereses_pagados = 0
        abonos_mes = 0

        historial = []

        for p in prestamos:

            prestamo_id = p[0]

            total_prestado += float(
                p[1]
            )

            # ====================================
            # 🔥 CALCULAR SALDO
            # ====================================
            cap_rest, int_rest, saldo, _, _ = calcular(
                prestamo_id,
                conn
            )

            saldo_total += saldo

            # ====================================
            # 💳 ABONOS DEL MES
            # ====================================
            cur.execute("""
                SELECT
                    monto,
                    tipo,
                    fecha
                FROM abonos
                WHERE prestamo_id=%s
                AND fecha >= %s
                AND fecha < %s
                ORDER BY fecha DESC
            """, (
                prestamo_id,
                inicio_mes,
                siguiente
            ))

            abonos = cur.fetchall()

            for a in abonos:

                abonos_mes += float(a[0])

                if a[1] == "interes":

                    intereses_pagados += float(
                        a[0]
                    )

            historial.append({

                "prestamo": p,
                "abonos": abonos

            })

        reporte.append({

            "cliente": cliente,
            "prestado": total_prestado,
            "saldo": saldo_total,
            "intereses": intereses_pagados,
            "abonos_mes": abonos_mes,
            "historial": historial

        })

    conn.close()

    return render_template(

        "reporte_general.html",

        reporte=reporte,

        formato=formato,

        mes=fecha.strftime("%Y-%m")

    )

# ====================================
# 🔄 CAMBIAR ESTADO PRÉSTAMO
# ====================================
@app.route("/cambiar_estado/<int:id>")
def cambiar_estado(id):

    conn = conectar()
    cur = conn.cursor()

    try:

        # 🔍 ESTADO ACTUAL
        cur.execute("""
            SELECT estado
            FROM prestamos
            WHERE id=%s
        """, (id,))

        data = cur.fetchone()

        if data:

            estado_actual = data[0]

            nuevo_estado = "Inactivo"

            if estado_actual == "Inactivo":
                nuevo_estado = "Activo"

            # 🔥 ACTUALIZAR
            cur.execute("""
                UPDATE prestamos
                SET estado=%s
                WHERE id=%s
            """, (

                nuevo_estado,
                id

            ))

            conn.commit()

    except Exception as e:

        print("ERROR CAMBIANDO ESTADO:", e)

    conn.close()

    return redirect(url_for("reporte_general"))

# ✏️ EDITAR MOVIMIENTO CAJA
# ====================================
@app.route("/editar_caja/<int:id>", methods=["GET", "POST"])
def editar_caja(id):

    conn = conectar()

    if not conn:
        return "Error conexión DB"

    cur = conn.cursor()

    error = ""

    # ====================================
    # 💾 GUARDAR CAMBIOS
    # ====================================
    if request.method == "POST":

        try:

            tipo = request.form["tipo"]

            monto = float(
                request.form["monto"]
                .replace(".", "")
                .replace(",", "")
            )

            descripcion = request.form["descripcion"]

            # ====================================
            # 🔥 OBTENER MOVIMIENTO ACTUAL
            # ====================================
            cur.execute("""
                SELECT
                    tipo,
                    monto
                FROM caja
                WHERE id=%s
            """, (id,))

            actual = cur.fetchone()

            # 🔥 SI NO EXISTE
            if not actual:

                conn.close()

                return redirect(
                    url_for("caja")
                )

            tipo_actual = actual[0]

            monto_actual = float(
                actual[1]
            )

            # ====================================
            # 💰 TOTALES ACTUALES
            # ====================================
            cur.execute("""
                SELECT

                    COALESCE(SUM(
                        CASE
                            WHEN tipo='ingreso'
                            THEN monto
                            ELSE 0
                        END
                    ),0),

                    COALESCE(SUM(
                        CASE
                            WHEN tipo='egreso'
                            THEN monto
                            ELSE 0
                        END
                    ),0)

                FROM caja
            """)

            datos = cur.fetchone()

            ingresos = float(
                datos[0] or 0
            )

            egresos = float(
                datos[1] or 0
            )

            # ====================================
            # 🔥 QUITAR MOVIMIENTO ACTUAL
            # ====================================
            if tipo_actual == "ingreso":

                ingresos -= monto_actual

            else:

                egresos -= monto_actual

            # ====================================
            # 🔥 AGREGAR NUEVO MOVIMIENTO
            # ====================================
            if tipo == "ingreso":

                ingresos += monto

            else:

                egresos += monto

            # ====================================
            # 💰 DISPONIBLE
            # ====================================
            disponible = (
                ingresos - egresos
            )

            # ====================================
            # ❌ VALIDAR NEGATIVO
            # ====================================
            if disponible < 0:

                error = (
                    "❌ Esa edición dejaría "
                    "la caja en negativo"
                )

            else:

                # ====================================
                # 💾 ACTUALIZAR
                # ====================================
                cur.execute("""
                    UPDATE caja
                    SET

                        tipo=%s,
                        monto=%s,
                        descripcion=%s

                    WHERE id=%s
                """, (

                    tipo,
                    monto,
                    descripcion,
                    id

                ))

                conn.commit()

                conn.close()

                return redirect(
                    url_for("caja")
                )

        except Exception as e:

            print(
                "ERROR EDITAR CAJA:",
                e
            )

            error = (
                "❌ Error editando movimiento"
            )

    # ====================================
    # 🔍 OBTENER MOVIMIENTO
    # ====================================
    cur.execute("""
        SELECT

            id,
            tipo,
            monto,
            descripcion,
            fecha

        FROM caja

        WHERE id=%s
    """, (id,))

    movimiento = cur.fetchone()

    conn.close()

    return render_template(

        "editar_caja.html",

        movimiento=movimiento,

        formato=formato,

        error=error

    )
@app.route("/eliminar_caja/<int:id>")
def eliminar_caja(id):

    conn = conectar()
    cur = conn.cursor()

    try:

        # ====================================
        # 🔥 OBTENER MOVIMIENTO
        # ====================================
        cur.execute("""
            SELECT
                tipo,
                monto
            FROM caja
            WHERE id=%s
        """, (id,))

        movimiento = cur.fetchone()

        if not movimiento:

            conn.close()

            return redirect(
                url_for("caja")
            )

        tipo = movimiento[0]
        monto = float(movimiento[1])

        # ====================================
        # 💰 VALIDAR SI SE PUEDE ELIMINAR
        # ====================================
        cur.execute("""
            SELECT

                COALESCE(SUM(
                    CASE
                        WHEN tipo='ingreso'
                        THEN monto
                        ELSE 0
                    END
                ),0),

                COALESCE(SUM(
                    CASE
                        WHEN tipo='egreso'
                        THEN monto
                        ELSE 0
                    END
                ),0)

            FROM caja
        """)

        datos = cur.fetchone()

        ingresos = float(datos[0] or 0)
        egresos = float(datos[1] or 0)

        disponible = ingresos - egresos

        # ====================================
        # ❌ NO DEJAR NEGATIVO
        # ====================================
        if (
            tipo == "ingreso"
            and (disponible - monto) < 0
        ):

            conn.close()

            return (
                "❌ No puedes eliminar este ingreso "
                "porque dejaría la caja en negativo"
            )

        # ====================================
        # 🗑️ ELIMINAR
        # ====================================
        cur.execute("""
            DELETE FROM caja
            WHERE id=%s
        """, (id,))

        conn.commit()

    except Exception as e:

        print("ERROR ELIMINAR CAJA:", e)

    conn.close()

    return redirect(
        url_for("caja")
    )

def crear_tabla_clientes():
    conn = conectar()

    if not conn:
        return

    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS clientes (
            id SERIAL PRIMARY KEY,
            nombre VARCHAR(200) NOT NULL,
            telefono VARCHAR(50),
            direccion VARCHAR(255),
            usuario VARCHAR(100)
        )
    """)

    conn.commit()
    conn.close()

    
# ------------------------------
# 👥 CLIENTES
# ------------------------------
@app.route("/clientes", methods=["GET","POST"])
def clientes():

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        cur.execute(
            "INSERT INTO clientes(nombre,telefono,direccion,usuario) VALUES (%s,%s,%s,%s)",
            (
                request.form["nombre"],
                request.form["telefono"],
                request.form["direccion"],
                session.get("usuario")
            )
        )
        conn.commit()

    # 🔹 Mostrar clientes
    cur.execute("SELECT * FROM clientes")
    clientes = cur.fetchall()

    resumen = []
    for c in clientes:
        cur.execute("SELECT SUM(capital) FROM prestamos WHERE cliente_id=%s", (c[0],))
        total = cur.fetchone()[0] or 0
        resumen.append(total)

    conn.close()

    return render_template(
        "clientes.html",
        clientes=clientes,
        resumen=resumen,
        formato=formato
    )
# ------------------------------
# 🔐 LOGIN
# ------------------------------

def actualizar_tabla_clientes():

    conn = conectar()

    if not conn:
        return

    cur = conn.cursor()

    try:

        cur.execute("""
            ALTER TABLE clientes
            ADD COLUMN IF NOT EXISTS telefono VARCHAR(50)
        """)

        cur.execute("""
            ALTER TABLE clientes
            ADD COLUMN IF NOT EXISTS direccion VARCHAR(255)
        """)

        cur.execute("""
            ALTER TABLE clientes
            ADD COLUMN IF NOT EXISTS usuario VARCHAR(100)
        """)

        conn.commit()

    except Exception as e:
        print("Error actualizando clientes:", e)

    finally:
        conn.close()

@app.route("/login", methods=["GET", "POST"])
def login():

    conn = conectar()

    if not conn:
        return "Error de conexión DB"

    cur = conn.cursor()

    if request.method == "POST":

        user = request.form["username"]
        password = request.form["password"]

        cur.execute("""
            SELECT * FROM usuarios
            WHERE username=%s AND password=%s
        """, (user, password))

        usuario = cur.fetchone()

        if usuario:

            session["usuario"] = user

            return redirect(url_for("panel"))

        else:

            return render_template(
                "login.html",
                error="Credenciales incorrectas"
            )

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))
# ------------------------------
# ✏️ EDITAR CLIENTE
# ------------------------------
@app.route("/editar_cliente/<int:id>", methods=["GET","POST"])
def editar_cliente(id):
    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        cur.execute("""
            UPDATE clientes
            SET nombre=%s, telefono=%s, direccion=%s
            WHERE id=%s
        """, (
            request.form["nombre"],
            request.form["telefono"],
            request.form["direccion"],
            id
        ))
        conn.commit()
        conn.close()
        return redirect(url_for("clientes"))

    cur.execute("SELECT * FROM clientes WHERE id=%s", (id,))
    cliente = cur.fetchone()

    conn.close()
    return render_template("editar_cliente.html", cliente=cliente)

@app.route("/guardar_abono", methods=["POST"])
def guardar_abono():
    try:
        prestamo_id = request.form.get("prestamo_id")
        meses = int(request.form.get("meses"))
        interes_mensual = float(request.form.get("interes_mensual"))

        monto = meses * interes_mensual

        conn = conectar()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO abonos (prestamo_id, monto, tipo, fecha)
            VALUES (%s, %s, %s, NOW())
        """, (prestamo_id, monto, "interes"))

        conn.commit()
        conn.close()

        return {"ok": True}

    except Exception as e:
        print("ERROR GUARDAR ABONO:", e)
        return {"ok": False}
    
# ------------------------------
# 🗑 ELIMINAR CLIENTE
# ------------------------------
@app.route("/eliminar_cliente/<int:id>")
def eliminar_cliente(id):
    conn = conectar()
    cur = conn.cursor()

    cur.execute("DELETE FROM abonos WHERE prestamo_id IN (SELECT id FROM prestamos WHERE cliente_id=%s)", (id,))
    cur.execute("DELETE FROM prestamos WHERE cliente_id=%s", (id,))
    cur.execute("DELETE FROM clientes WHERE id=%s", (id,))

    conn.commit()
    conn.close()

    return redirect(url_for("clientes"))

# ------------------------------
# 📄 HISTORIAL
# ------------------------------
@app.route("/historial/<int:id>")
def historial(id):

    conn = conectar()
    cur = conn.cursor()

    cur.execute(
        "SELECT nombre FROM clientes WHERE id=%s",
        (id,)
    )

    cliente = cur.fetchone()

    cur.execute("""
        SELECT id, capital, total, fecha
        FROM prestamos
        WHERE cliente_id=%s
        ORDER BY id DESC
    """, (id,))

    prestamos = cur.fetchall()

    historial = []

    for p in prestamos:

        cur.execute("""
            SELECT monto, tipo, fecha
            FROM abonos
            WHERE prestamo_id=%s
            ORDER BY fecha DESC
        """, (p[0],))

        abonos = cur.fetchall()

        historial.append({
            "prestamo": p,
            "abonos": abonos
        })

    conn.close()

    return render_template(
        "historial.html",
        cliente=cliente,
        historial=historial,
        formato=formato
    )

def crear_tabla_prestamos():
    conn = conectar()
    if not conn:
        return

    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS prestamos (
            id SERIAL PRIMARY KEY,
            cliente_id INTEGER NOT NULL,
            capital NUMERIC(15,2) NOT NULL,
            interes NUMERIC(10,2) NOT NULL,
            fecha DATE NOT NULL,
            vencimiento DATE NOT NULL,
            total NUMERIC(15,2) NOT NULL,
            estado VARCHAR(30) DEFAULT 'Activo'
        )
    """)

    conn.commit()
    conn.close()


# 💰 PRÉSTAMOS
# ====================================
@app.route("/prestamos", methods=["GET","POST"])
def prestamos():

    conn = conectar()

    if not conn:
        return "Error conexión DB"

    cur = conn.cursor()

    error = ""

    # ====================================
    # 👥 CLIENTES
    # ====================================
    cur.execute("""
        SELECT *
        FROM clientes
        ORDER BY nombre ASC
    """)

    clientes = cur.fetchall()

    # ====================================
    # 💾 GUARDAR PRÉSTAMO
    # ====================================
    if request.method == "POST":

        try:

            capital = float(
                request.form["capital"]
                .replace(".", "")
                .replace(",", "")
            )

            interes = float(
                request.form["interes"]
            )

            fecha = datetime.strptime(
                request.form["fecha"],
                "%Y-%m-%d"
            )

            venc = datetime.strptime(
                request.form["vencimiento"],
                "%Y-%m-%d"
            )

            # ====================================
            # ❌ VALIDAR FECHA
            # ====================================
            if venc <= fecha:

                error = (
                    "❌ La fecha de vencimiento debe ser mayor"
                )

            else:

                # ====================================
                # 💰 DINERO DISPONIBLE EN CAJA
                # ====================================
                cur.execute("""
                    SELECT

                        COALESCE(SUM(
                            CASE
                                WHEN tipo='ingreso'
                                THEN monto
                                ELSE 0
                            END
                        ),0),

                        COALESCE(SUM(
                            CASE
                                WHEN tipo='egreso'
                                THEN monto
                                ELSE 0
                            END
                        ),0)

                    FROM caja
                """)

                datos = cur.fetchone()

                ingresos = float(
                    datos[0] or 0
                )

                egresos = float(
                    datos[1] or 0
                )

                disponible = (
                    ingresos - egresos
                )

                # ====================================
                # ❌ VALIDAR DINERO EN CAJA
                # ====================================
                if capital > disponible:

                    error = (
                        f"❌ No es posible realizar el préstamo. "
                        f"Dinero insuficiente en caja. "
                        f"Disponible: ${formato(disponible)}"
                    )

                else:

                    total = capital + (
                        capital * interes / 100
                    )

                    # ====================================
                    # 💾 GUARDAR PRÉSTAMO
                    # ====================================
                    cur.execute("""
                        INSERT INTO prestamos (

                            cliente_id,
                            capital,
                            interes,
                            fecha,
                            vencimiento,
                            total,
                            estado

                        )
                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                    """, (

                        request.form["cliente"],
                        capital,
                        interes,
                        fecha.date(),
                        venc.date(),
                        total,
                        "Activo"

                    ))

                    # ====================================
                    # 🏦 EGRESO EN CAJA
                    # ====================================
                    cur.execute("""
                        INSERT INTO caja (

                            tipo,
                            monto,
                            descripcion

                        )
                        VALUES (%s,%s,%s)
                    """, (

                        "egreso",
                        capital,
                        "Préstamo entregado"

                    ))

                    conn.commit()

        except Exception as e:

            print("ERROR PRESTAMOS:", e)

            error = (
                "❌ Error al guardar préstamo"
            )

    # ====================================
    # 🔍 FILTRO POR FECHA
    # ====================================
    fecha_filtro = request.args.get("fecha")

    if fecha_filtro:

        fecha_filtro = datetime.strptime(
            fecha_filtro,
            "%Y-%m-%d"
        ).date()

    else:

        fecha_filtro = datetime.now().date()

    # ====================================
    # 📋 PRÉSTAMOS DEL DÍA
    # ====================================
    cur.execute("""
        SELECT

            p.id,
            c.nombre,
            p.capital,
            p.total,
            p.fecha

        FROM prestamos p

        JOIN clientes c
        ON p.cliente_id = c.id

        WHERE p.fecha = %s
        AND p.estado = 'Activo'

        ORDER BY p.id DESC

    """, (fecha_filtro,))

    prestamos_dia = cur.fetchall()

    cantidad_dia = len(
        prestamos_dia
    )

    # ====================================
    # 📋 LISTA GENERAL
    # ====================================
    cur.execute("""
        SELECT

            p.id,
            c.nombre,
            p.capital,
            p.interes,
            p.fecha,
            p.estado

        FROM prestamos p

        JOIN clientes c
        ON p.cliente_id = c.id

        WHERE p.estado = 'Activo'

        ORDER BY p.id DESC
    """)

    prestamos_lista = []

    resultados = cur.fetchall()

    for p in resultados:

        try:

            pid = p[0]
            nombre = p[1]
            estado = p[5]

            # ====================================
            # 🔥 CÁLCULO REAL
            # ====================================
            cap_rest, int_rest, saldo_total, _, _ = calcular(
                pid,
                conn
            )

            # ====================================
            # 🔥 SI YA PAGÓ TODO
            # ====================================
            if saldo_total <= 0:
                continue

            prestamos_lista.append({

                "id": pid,

                "cliente": nombre,

                "estado": estado,

                "total": formato(
                    saldo_total
                ),

                "saldo": formato(
                    saldo_total
                )

            })

        except Exception as e:

            print(
                "ERROR CALCULANDO:",
                e
            )

    conn.close()

    return render_template(

        "prestamos.html",

        clientes=clientes,

        prestamos=prestamos_lista,

        prestamos_dia=prestamos_dia,

        cantidad_dia=cantidad_dia,

        fecha=fecha_filtro,

        error=error

    )
    
def crear_tabla_abonos():
    conn = conectar()

    if not conn:
        return

    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS abonos (
            id SERIAL PRIMARY KEY,
            prestamo_id INTEGER NOT NULL,
            monto NUMERIC(15,2) NOT NULL,
            fecha TIMESTAMP NOT NULL,
            tipo VARCHAR(30) NOT NULL,
            mes INTEGER DEFAULT 0
        )
    """)

    conn.commit()
    conn.close()

# 💸 ABONOS
# ------------------------------
@app.route("/abonos", methods=["GET", "POST"])
def abonos():

    conn = conectar()

    if not conn:
        return "Error conexión DB"

    cur = conn.cursor()

    # =============================
    # 👥 CLIENTES
    # =============================
    cur.execute("SELECT * FROM clientes")
    clientes = cur.fetchall()

    prestamos = []
    mensaje = ""

    cliente_id = (
        request.form.get("cliente")
        or request.args.get("cliente")
    )

    # =============================
    # 🔍 CARGAR PRÉSTAMOS
    # =============================
    if cliente_valido(cliente_id):

        try:

            cur.execute("""
                SELECT
                    p.id,
                    c.nombre,
                    p.fecha
                FROM prestamos p
                JOIN clientes c
                ON p.cliente_id = c.id
                WHERE c.id = %s
            """, (cliente_id,))

            resultados = cur.fetchall()

            for fila in resultados:

                pid = fila[0]
                nombre = fila[1]
                fecha_prestamo = fila[2]

                # =============================
                # 💰 CALCULAR SALDOS
                # =============================
                cap_rest, int_rest, total, _, _ = calcular(
                    pid,
                    conn
                )

                cap_rest = round(cap_rest)
                int_rest = round(int_rest)
                total = round(total)

                if total <= 0:
                    continue

                # =============================
                # 📈 INTERÉS MENSUAL
                # =============================
                cur.execute("""
                    SELECT capital, interes
                    FROM prestamos
                    WHERE id=%s
                """, (pid,))

                data = cur.fetchone()

                if data:

                    capital_original = float(data[0])
                    tasa = float(data[1])

                    interes_mensual = round(
                        capital_original *
                        (tasa / 100)
                    )

                else:

                    interes_mensual = 0

                # =============================
                # 📅 MESES TRANSCURRIDOS
                # =============================
                meses_transcurridos = meses_atraso(
                    fecha_prestamo
                )

                if meses_transcurridos < 1:
                    meses_transcurridos = 1

                # 🔥 PERMITIR ADELANTAR
                meses_totales = (
                    meses_transcurridos + 24
                )

                # =============================
                # 📌 MESES PAGADOS
                # =============================
                cur.execute("""
                    SELECT DISTINCT mes
                    FROM abonos
                    WHERE prestamo_id=%s
                    AND tipo='interes'
                    AND mes IS NOT NULL
                    ORDER BY mes ASC
                """, (pid,))

                meses_pagados_db = [

                    int(x[0])

                    for x in cur.fetchall()

                    if x[0] is not None

                ]

                # =============================
                # 📌 ÚLTIMO MES PAGADO
                # =============================
                ultimo_mes_pagado = "Sin pagos"

                if meses_pagados_db:

                    ultimo_num = max(
                        meses_pagados_db
                    )

                    fecha_ultimo = (
                        fecha_prestamo +
                        relativedelta(
                            months=ultimo_num - 1
                        )
                    )

                    ultimo_mes_pagado = (
                        fecha_ultimo.strftime(
                            "%B %Y"
                        ).capitalize()
                    )

                # =============================
                # 📋 GENERAR MESES
                # 🔥 OCULTAR PAGADOS
                # =============================
                meses = []

                for i in range(
                    1,
                    meses_totales + 1
                ):

                    # 🔥 SI YA ESTÁ PAGADO
                    if i in meses_pagados_db:
                        continue

                    fecha_mes = (
                        fecha_prestamo +
                        relativedelta(
                            months=i - 1
                        )
                    )

                    texto_mes = (
                        fecha_mes.strftime(
                            "%B %Y"
                        ).capitalize()
                    )

                    meses.append({

                        "numero": i,
                        "texto": texto_mes

                    })

                # =============================
                # 📆 CANTIDAD MESES PAGADOS
                # =============================
                cantidad_meses = len(
                    meses_pagados_db
                )

                # =============================
                # ➕ AGREGAR
                # =============================
                prestamos.append({

                    "id": pid,
                    "nombre": nombre,

                    "capital":
                        formato(cap_rest),

                    "interes":
                        formato(int_rest),

                    "total":
                        formato(total),

                    "capital_num":
                        cap_rest,

                    "interes_num":
                        interes_mensual,

                    "total_num":
                        total,

                    "meses":
                        meses,

                    "interes_mensual":
                        formato(interes_mensual),

                    "interes_mensual_raw":
                        interes_mensual,

                    "ultimo_mes_pagado":
                        ultimo_mes_pagado,

                    "cantidad_meses":
                        cantidad_meses

                })

        except Exception as e:

            print(
                "Error cargando préstamos:",
                e
            )

    # =============================
    # 💾 GUARDAR ABONO
    # =============================
    if request.method == "POST":

        if not request.form.get("prestamo"):

            conn.close()

            return render_template(

                "abonos.html",

                clientes=clientes,
                prestamos=prestamos,
                mensaje=mensaje,
                cliente_id=cliente_id

            )

        try:

            pid = int(
                request.form.get(
                    "prestamo"
                )
            )

            monto = float(
                request.form.get(
                    "monto"
                )
                .replace(".", "")
                .replace(",", "")
            )

            tipo = request.form.get(
                "tipo"
            )

            # =============================
            # 📅 MES PAGADO
            # =============================
            if tipo == "capital":

                mes_pagado = 0

            else:

                mes_pagado = int(
                    request.form.get(
                        "mes"
                    ) or 0
                )

            # =============================
            # 💰 VALIDACIONES
            # =============================
            cap_rest, int_rest, _, _, _ = calcular(
                pid,
                conn
            )

            if (
                tipo == "capital"
                and monto > cap_rest
            ):

                mensaje = (
                    "❌ Excede capital"
                )

            elif (
                tipo == "interes"
                and monto <= 0
            ):

                mensaje = (
                    "❌ Valor inválido"
                )

            else:

                # =====================================
                # 🔥 ABONO INTERÉS
                # =====================================
                if tipo == "interes":

                    # 🔥 BUSCAR INTERÉS MENSUAL
                    cur.execute("""
                        SELECT capital, interes
                        FROM prestamos
                        WHERE id=%s
                    """, (pid,))

                    data = cur.fetchone()

                    capital_original = float(data[0])
                    tasa = float(data[1])

                    interes_mensual = round(
                        capital_original *
                        (tasa / 100)
                    )

                    total_ingreso = 0

                    # 🔥 GUARDAR TODOS LOS MESES
                    for mes_actual in range(
                        1,
                        mes_pagado + 1
                    ):

                        # 🔥 VALIDAR DUPLICADO
                        cur.execute("""
                            SELECT id
                            FROM abonos
                            WHERE prestamo_id=%s
                            AND tipo='interes'
                            AND mes=%s
                        """, (

                            pid,
                            mes_actual

                        ))

                        existe = cur.fetchone()

                        # 🔥 SI NO EXISTE
                        if not existe:

                            cur.execute("""
                                INSERT INTO abonos (

                                    prestamo_id,
                                    monto,
                                    fecha,
                                    tipo,
                                    mes

                                )
                                VALUES (

                                    %s,
                                    %s,
                                    %s,
                                    %s,
                                    %s

                                )
                            """, (

                                pid,
                                interes_mensual,
                                datetime.now(),
                                "interes",
                                mes_actual

                            ))

                            total_ingreso += interes_mensual

                    # =====================================
                    # 🏦 INGRESO EN CAJA
                    # =====================================
                    if total_ingreso > 0:

                        cur.execute("""
                            INSERT INTO caja (

                                tipo,
                                monto,
                                descripcion

                            )
                            VALUES (%s,%s,%s)
                        """, (

                            "ingreso",
                            total_ingreso,
                            "Abono interes"

                        ))

                else:

                    # =====================================
                    # 💰 ABONO CAPITAL
                    # =====================================
                    cur.execute("""
                        INSERT INTO abonos (

                            prestamo_id,
                            monto,
                            fecha,
                            tipo,
                            mes

                        )
                        VALUES (

                            %s,
                            %s,
                            %s,
                            %s,
                            %s

                        )
                    """, (

                        pid,
                        round(monto),
                        datetime.now(),
                        tipo,
                        mes_pagado

                    ))

                    # =====================================
                    # 🏦 INGRESO EN CAJA
                    # =====================================
                    cur.execute("""
                        INSERT INTO caja (

                            tipo,
                            monto,
                            descripcion

                        )
                        VALUES (%s,%s,%s)
                    """, (

                        "ingreso",
                        round(monto),
                        "Abono capital"

                    ))

                conn.commit()

                conn.close()

                return redirect(
                    url_for(
                        "abonos",
                        cliente=cliente_id
                    )
                )

        except Exception as e:

            print(
                "Error abonos:",
                e
            )

            mensaje = (
                "❌ Error en datos"
            )

    conn.close()

    return render_template(

        "abonos.html",

        clientes=clientes,
        prestamos=prestamos,
        mensaje=mensaje,
        cliente_id=cliente_id

    )

# ====================================
# 🗑 ELIMINAR ABONO
# ====================================
@app.route("/eliminar_abono/<int:id>")
def eliminar_abono(id):

    conn = conectar()
    cur = conn.cursor()

    try:

        # 🔥 obtener préstamo
        cur.execute("""
            SELECT prestamo_id
            FROM abonos
            WHERE id=%s
        """, (id,))

        data = cur.fetchone()

        if not data:

            conn.close()

            return redirect(url_for("clientes"))

        prestamo_id = data[0]

        # 🔥 obtener cliente
        cur.execute("""
            SELECT cliente_id
            FROM prestamos
            WHERE id=%s
        """, (prestamo_id,))

        cliente = cur.fetchone()

        cliente_id = cliente[0]

        # 🔥 eliminar abono
        cur.execute("""
            DELETE FROM abonos
            WHERE id=%s
        """, (id,))

        conn.commit()

        conn.close()

        return redirect(
            url_for(
                "historial",
                id=cliente_id
            )
        )

    except Exception as e:

        print("ERROR ELIMINAR ABONO:", e)

        conn.close()

        return "❌ Error eliminando abono"
# ------------------------------
# 🔥 OBTENER INTERÉS AUTOMÁTICO
# ------------------------------
@app.route("/obtener_interes")
def obtener_interes():

    prestamo_id = request.args.get("prestamo_id")
    mes = request.args.get("mes")

    conn = conectar()

    if not conn:
        return jsonify({"interes": 0})

    cur = conn.cursor()

    try:

        # =============================
        # VALIDAR SI YA PAGÓ ESE MES
        # =============================
        if mes:

            cur.execute("""
                SELECT COUNT(*)
                FROM abonos
                WHERE prestamo_id=%s
                AND tipo='interes'
                AND mes=%s
            """, (

                prestamo_id,
                mes

            ))

            ya_pagado = cur.fetchone()[0]

            if ya_pagado > 0:

                return jsonify({
                    "interes": 0
                })

        # =============================
        # OBTENER DATOS
        # =============================
        cur.execute("""
            SELECT capital, interes
            FROM prestamos
            WHERE id=%s
        """, (prestamo_id,))

        data = cur.fetchone()

        if data:

            capital = float(data[0])
            interes = float(data[1])

            interes_mensual = (
                capital *
                (interes / 100)
            )

        else:

            interes_mensual = 0

    except Exception as e:

        print("ERROR obtener_interes:", e)

        interes_mensual = 0

    finally:

        cur.close()
        conn.close()

    return jsonify({

        "interes":
        round(interes_mensual, 2)

    })

# ============================================
# 📊 REPORTE EXCEL PREMIUM
# ============================================

@app.route("/reporte_excel")
def reporte_excel():

    conn = conectar()

    if not conn:
        return "Error DB"

    cur = conn.cursor()

    # ============================================
    # 📅 FILTRO POR MES
    # ============================================

    mes = request.args.get("mes", "").strip()

    # 🔥 SI VIENE VACÍO O MAL FORMADO
    if not mes or "-" not in mes:

        mes = datetime.now().strftime("%Y-%m")

    try:

        anio = int(mes.split("-")[0])
        mes_num = int(mes.split("-")[1])

    except:

        anio = datetime.now().year
        mes_num = datetime.now().month

        mes = datetime.now().strftime("%Y-%m")

    # ============================================
    # 📦 CREAR EXCEL
    # ============================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Reporte General"

    # ============================================
    # 🎨 ESTILOS
    # ============================================

    azul = PatternFill(
        start_color="1E40AF",
        end_color="1E40AF",
        fill_type="solid"
    )

    naranja = PatternFill(
        start_color="F97316",
        end_color="F97316",
        fill_type="solid"
    )

    verde = PatternFill(
        start_color="22C55E",
        end_color="22C55E",
        fill_type="solid"
    )

    rojo = PatternFill(
        start_color="EF4444",
        end_color="EF4444",
        fill_type="solid"
    )

    blanco = Font(
        color="FFFFFF",
        bold=True
    )

    titulo = Font(
        size=16,
        bold=True
    )

    negrita = Font(
        bold=True
    )

    center = Alignment(
        horizontal="center"
    )

    borde = Border(

        left=Side(style='thin'),

        right=Side(style='thin'),

        top=Side(style='thin'),

        bottom=Side(style='thin')

    )

    # ============================================
    # 🔥 TÍTULO
    # ============================================

    ws.merge_cells("A1:H1")

    ws["A1"] = f"REPORTE GENERAL - {mes}"

    ws["A1"].font = titulo

    ws["A1"].alignment = center

    fila = 3

    # ============================================
    # 👥 TOTAL CLIENTES
    # ============================================

    cur.execute("""
        SELECT COUNT(*)
        FROM clientes
    """)

    total_clientes = cur.fetchone()[0]

    # ============================================
    # 💰 DINERO EN CAJA
    # ============================================

    cur.execute("""

        SELECT

            COALESCE(SUM(
                CASE
                    WHEN tipo='ingreso'
                    THEN monto
                    ELSE 0
                END
            ),0),

            COALESCE(SUM(
                CASE
                    WHEN tipo='egreso'
                    THEN monto
                    ELSE 0
                END
            ),0)

        FROM caja

    """)

    datos_caja = cur.fetchone()

    ingresos = float(datos_caja[0] or 0)

    egresos = float(datos_caja[1] or 0)

    disponible = ingresos - egresos

    # ============================================
    # 💸 TOTAL PRESTADO
    # ============================================

    cur.execute("""

        SELECT COALESCE(SUM(capital),0)

        FROM prestamos

        WHERE EXTRACT(YEAR FROM fecha)=%s
        AND EXTRACT(MONTH FROM fecha)=%s

    """, (

        anio,
        mes_num

    ))

    total_prestado = float(
        cur.fetchone()[0] or 0
    )

    # ============================================
    # 📊 MÉTRICAS
    # ============================================

    datos = [

        ["Total clientes", total_clientes],

        ["Ingresos", ingresos],

        ["Egresos", egresos],

        ["Disponible en caja", disponible],

        ["Total prestado mes", total_prestado]

    ]

    for d in datos:

        ws[f"A{fila}"] = d[0]
        ws[f"B{fila}"] = d[1]

        ws[f"A{fila}"].font = negrita
        ws[f"A{fila}"].border = borde
        ws[f"B{fila}"].border = borde

        fila += 1

    fila += 2

    # ============================================
    # 👥 CLIENTES
    # ============================================

    ws.merge_cells(
        start_row=fila,
        start_column=1,
        end_row=fila,
        end_column=5
    )

    ws.cell(
        row=fila,
        column=1
    ).value = "CLIENTES"

    ws.cell(
        row=fila,
        column=1
    ).fill = azul

    ws.cell(
        row=fila,
        column=1
    ).font = blanco

    ws.cell(
        row=fila,
        column=1
    ).alignment = center

    fila += 1

    encabezados = [

        "ID",
        "Nombre",
        "Teléfono",
        "Dirección",
        "Usuario"

    ]

    for col, valor in enumerate(encabezados, 1):

        c = ws.cell(
            row=fila,
            column=col
        )

        c.value = valor

        c.fill = naranja

        c.font = blanco

        c.alignment = center

        c.border = borde

    fila += 1

    cur.execute("""

        SELECT
            id,
            nombre,
            telefono,
            direccion,
            usuario

        FROM clientes

        ORDER BY nombre ASC

    """)

    clientes = cur.fetchall()

    for c in clientes:

        for col, valor in enumerate(c, 1):

            celda = ws.cell(
                row=fila,
                column=col
            )

            celda.value = valor

            celda.border = borde

        fila += 1

    fila += 3

    # ============================================
    # 💸 PRÉSTAMOS
    # ============================================

    ws.merge_cells(
        start_row=fila,
        start_column=1,
        end_row=fila,
        end_column=8
    )

    ws.cell(
        row=fila,
        column=1
    ).value = "PRÉSTAMOS"

    ws.cell(
        row=fila,
        column=1
    ).fill = verde

    ws.cell(
        row=fila,
        column=1
    ).font = blanco

    ws.cell(
        row=fila,
        column=1
    ).alignment = center

    fila += 1

    encabezados = [

        "ID",
        "Cliente",
        "Capital",
        "Interés",
        "Fecha",
        "Vencimiento",
        "Total",
        "Estado"

    ]

    for col, valor in enumerate(encabezados, 1):

        c = ws.cell(
            row=fila,
            column=col
        )

        c.value = valor

        c.fill = naranja

        c.font = blanco

        c.alignment = center

        c.border = borde

    fila += 1

    cur.execute("""

        SELECT

            p.id,
            c.nombre,
            p.capital,
            p.interes,
            p.fecha,
            p.vencimiento,
            p.total,
            p.estado

        FROM prestamos p

        JOIN clientes c
        ON p.cliente_id = c.id

        WHERE EXTRACT(YEAR FROM p.fecha)=%s
        AND EXTRACT(MONTH FROM p.fecha)=%s

        ORDER BY p.id DESC

    """, (

        anio,
        mes_num

    ))

    prestamos = cur.fetchall()

    for p in prestamos:

        for col, valor in enumerate(p, 1):

            celda = ws.cell(
                row=fila,
                column=col
            )

            celda.value = valor

            celda.border = borde

        fila += 1

    fila += 3

    # ============================================
    # 💵 ABONOS
    # ============================================

    ws.merge_cells(
        start_row=fila,
        start_column=1,
        end_row=fila,
        end_column=6
    )

    ws.cell(
        row=fila,
        column=1
    ).value = "ABONOS"

    ws.cell(
        row=fila,
        column=1
    ).fill = rojo

    ws.cell(
        row=fila,
        column=1
    ).font = blanco

    ws.cell(
        row=fila,
        column=1
    ).alignment = center

    fila += 1

    encabezados = [

        "ID",
        "Cliente",
        "Monto",
        "Fecha",
        "Tipo",
        "Usuario"

    ]

    for col, valor in enumerate(encabezados, 1):

        c = ws.cell(
            row=fila,
            column=col
        )

        c.value = valor

        c.fill = naranja

        c.font = blanco

        c.alignment = center

        c.border = borde

    fila += 1

    cur.execute("""

        SELECT

            a.id,
            c.nombre,
            a.monto,
            a.fecha,
            a.tipo,
            a.usuario

        FROM abonos a

        JOIN prestamos p
        ON a.prestamo_id = p.id

        JOIN clientes c
        ON p.cliente_id = c.id

        WHERE EXTRACT(YEAR FROM a.fecha)=%s
        AND EXTRACT(MONTH FROM a.fecha)=%s

        ORDER BY a.id DESC

    """, (

        anio,
        mes_num

    ))

    abonos = cur.fetchall()

    for a in abonos:

        for col, valor in enumerate(a, 1):

            celda = ws.cell(
                row=fila,
                column=col
            )

            celda.value = valor

            celda.border = borde

        fila += 1

    # ============================================
    # 📏 ANCHO COLUMNAS
    # ============================================

    columnas = [

        "A","B","C","D",
        "E","F","G","H"

    ]

    for col in columnas:

        ws.column_dimensions[col].width = 24

    # ============================================
    # 💾 GUARDAR MEMORIA
    # ============================================

    archivo = io.BytesIO()

    wb.save(archivo)

    archivo.seek(0)

    conn.close()

    # ============================================
    # ⬇ DESCARGAR
    # ============================================

    return send_file(

        archivo,

        as_attachment=True,

        download_name=f"reporte_{mes}.xlsx",

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

crear_admin()
crear_tabla_prestamos()
crear_tabla_abonos()
crear_tabla_clientes()
actualizar_tabla_clientes()
crear_tabla_caja()
if __name__ == "__main__":
    init_db()
    crear_admin()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)